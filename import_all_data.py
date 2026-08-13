import sqlite3
import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

db_path = r'c:\xampp\htdocs\Quran Center\QuranCircles.Api\QuranCircles.Api\quran.db'
excel_path = r'c:\xampp\htdocs\Quran Center\all data.xlsx'

conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Ensure ProfileUpdateRequests table exists
cursor.execute("""
CREATE TABLE IF NOT EXISTS ProfileUpdateRequests (
    Id INTEGER PRIMARY KEY AUTOINCREMENT,
    UserId INTEGER NOT NULL,
    StudentId INTEGER,
    RequestedByRole TEXT NOT NULL,
    RequestedByName TEXT NOT NULL,
    ChangesJson TEXT NOT NULL,
    Status TEXT NOT NULL DEFAULT 'Pending',
    RequestDate TEXT NOT NULL,
    ReviewerNotes TEXT,
    ReviewDate TEXT
);
""")

# Ensure Students table columns exist
cursor.execute("PRAGMA table_info('Students');")
existing_cols = [c[1] for c in cursor.fetchall()]

new_columns = {
    'StudentIdentityNumber': 'TEXT',
    'PreviousQuranMemorization': 'TEXT',
    'StudentMobile': 'TEXT',
    'StudentWhatsapp': 'TEXT',
    'HealthStatus': 'TEXT',
    'FatherStatus': 'TEXT',
    'MotherStatus': 'TEXT',
    'Kinship': 'TEXT',
    'ParentIdentityNumber': 'TEXT',
    'WhatsappNumber': 'TEXT',
    'WalletNumber': 'TEXT',
    'BankAccountNumber': 'TEXT',
    'BankName': 'TEXT',
    'OriginalAddress': 'TEXT',
    'OriginalHousingType': 'TEXT',
    'OriginalHousingStatus': 'TEXT',
    'CurrentAddress': 'TEXT',
    'CurrentHousingType': 'TEXT',
    'Notes': 'TEXT'
}

for col_name, col_type in new_columns.items():
    if col_name not in existing_cols:
        cursor.execute(f"ALTER TABLE Students ADD COLUMN {col_name} {col_type};")

conn.commit()

# 1. WIPE OLD STUDENT & PARENT DATA
print("1. Wiping old students, parents, sessions, and attendance data...")
cursor.execute("DELETE FROM Attendances;")
cursor.execute("DELETE FROM Sessions;")
cursor.execute("DELETE FROM HadithSessions;")
cursor.execute("DELETE FROM CourseAttendances;")
cursor.execute("DELETE FROM CourseEnrollments;")
cursor.execute("DELETE FROM ExamNominations;")
cursor.execute("DELETE FROM Students;")
cursor.execute("DELETE FROM ProfileUpdateRequests;")

# Remove old Student (Role = 3) and Parent (Role = 4) user login accounts
cursor.execute("DELETE FROM Users WHERE Role IN (3, 4);")
conn.commit()

# 2. READ EXCEL SHEET 'all data'
print("2. Reading sheet 'all data' from Excel...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['all data'] if 'all data' in wb.sheetnames else wb.active
rows = list(ws.iter_rows(values_only=True))[1:]

student_rows = [r for r in rows if r[1] and str(r[1]).strip()]
print(f"Found {len(student_rows)} valid student rows in Excel.")

# 3. CREATE/MAP CIRCLES
cursor.execute("SELECT Id, Name FROM Circles;")
existing_circles = {row[1]: row[0] for row in cursor.fetchall()}

circle_names_in_excel = set()
for r in student_rows:
    c_name = str(r[4]).strip() if r[4] and str(r[4]).strip() else "غير مسند حلقة"
    circle_names_in_excel.add(c_name)

circle_id_map = {}
for c_name in circle_names_in_excel:
    if c_name in existing_circles:
        circle_id_map[c_name] = existing_circles[c_name]
    else:
        cursor.execute("INSERT INTO Circles (Name, Timing, IsActive, TeacherId) VALUES (?, 1, 1, 1);", (c_name,))
        circle_id_map[c_name] = cursor.lastrowid

conn.commit()
print(f"Circles ready: {len(circle_id_map)} circles mapped.")

# 4. GROUP FAMILIES & CREATE PARENT ACCOUNTS (Username = ID number directly without prefix!)
family_groups = {}

for r in student_rows:
    p_id = str(r[14]).strip() if r[14] and str(r[14]).strip() else ""
    p_mobile = str(r[15]).strip() if r[15] and str(r[15]).strip() else ""
    p_name = str(r[12]).strip() if r[12] and str(r[12]).strip() else ""
    
    key = p_id or p_mobile or p_name
    if not key:
        key = f"unk_{len(family_groups)+1}"
        
    if key not in family_groups:
        family_groups[key] = {
            'parent_name': p_name or f"ولي أمر {r[1]}",
            'parent_id_num': p_id,
            'parent_mobile': p_mobile,
            'students': []
        }
    family_groups[key]['students'].append(r)

print(f"Total Unique Families (Parent accounts) to create: {len(family_groups)}")

parent_account_map = {}
for key, f_data in family_groups.items():
    p_name = f_data['parent_name']
    p_id_num = f_data['parent_id_num']
    p_mobile = f_data['parent_mobile']
    
    # USERNAME = ID NUMBER DIRECTLY WITHOUT PREFIX!
    clean_username = re.sub(r'\D', '', p_id_num or p_mobile or key)
    if not clean_username:
        clean_username = f"parent_{len(parent_account_map) + 1}"
        
    cursor.execute("SELECT Id FROM Users WHERE Username = ?;", (clean_username,))
    if cursor.fetchone():
        clean_username = f"{clean_username}_{len(parent_account_map)}"
        
    plain_pass = "123456"
    pass_hash = "AQAAAAIAAYagAAAAEP4qOslLTFhJOkkBoFWNLPRj6QTbtdjPKCf+rxQ/YWVpdIgXBcSHPb6td4ysp1WmUA=="

    cursor.execute("""
        INSERT INTO Users (Username, PasswordHash, PlainPassword, Role, FullName, IsActive, ParentId)
        VALUES (?, ?, ?, 4, ?, 1, NULL);
    """, (clean_username, pass_hash, plain_pass, p_name))
    
    user_id = cursor.lastrowid
    cursor.execute("UPDATE Users SET ParentId = ? WHERE Id = ?;", (user_id, user_id))
    parent_account_map[key] = user_id

conn.commit()
print("Parent accounts created with usernames set directly to ID numbers.")

# 5. CREATE STUDENTS WITH ALL 26 COLUMNS
inserted_count = 0
for key, f_data in family_groups.items():
    p_user_id = parent_account_map[key]
    
    for r in f_data['students']:
        st_name = str(r[1]).strip()
        st_id_num = str(r[2]).strip() if r[2] and str(r[2]).strip() else ""
        dob_raw = str(r[3]).strip() if r[3] and str(r[3]).strip() else ""
        
        dob = "2010-01-01"
        if dob_raw:
            parts = dob_raw.split('/')
            if len(parts) == 3:
                day, month, year = parts[0].zfill(2), parts[1].zfill(2), parts[2]
                dob = f"{year}-{month}-{day}"
                
        circle_name = str(r[4]).strip() if r[4] and str(r[4]).strip() else "غير مسند حلقة"
        circle_id = circle_id_map.get(circle_name)
        
        prev_quran = str(r[5]).strip() if r[5] else ""
        st_mobile = str(r[6]).strip() if r[6] else ""
        st_whatsapp = str(r[7]).strip() if r[7] else ""
        health_status = str(r[8]).strip() if r[8] else ""
        father_status = str(r[9]).strip() if r[9] else ""
        mother_status = str(r[10]).strip() if r[10] else ""
        kinship = str(r[13]).strip() if r[13] else ""
        p_id_num = str(r[14]).strip() if r[14] else ""
        family_contact = f_data['parent_mobile'] or str(r[15] or '')
        whatsapp_num = str(r[16]).strip() if r[16] else ""
        wallet_num = str(r[17]).strip() if r[17] else ""
        bank_acc = str(r[18]).strip() if r[18] else ""
        bank_name = str(r[19]).strip() if r[19] else ""
        orig_addr = str(r[20]).strip() if r[20] else ""
        orig_type = str(r[21]).strip() if r[21] else ""
        orig_status = str(r[22]).strip() if r[22] else ""
        curr_addr = str(r[23]).strip() if r[23] else ""
        curr_type = str(r[24]).strip() if r[24] else ""
        notes = str(r[25]).strip() if r[25] else ""

        address = curr_addr or orig_addr or 'غزة'

        cursor.execute("""
            INSERT INTO Students (
                FullName, Address, DateOfBirth, FamilyContact, RegistrationDate, IsActive, CircleId, ParentId,
                StudentIdentityNumber, PreviousQuranMemorization, StudentMobile, StudentWhatsapp, HealthStatus,
                FatherStatus, MotherStatus, Kinship, ParentIdentityNumber, WhatsappNumber, WalletNumber,
                BankAccountNumber, BankName, OriginalAddress, OriginalHousingType, OriginalHousingStatus,
                CurrentAddress, CurrentHousingType, Notes
            )
            VALUES (?, ?, ?, ?, '2026-08-05', 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, (
            st_name, address, dob, family_contact, circle_id, p_user_id,
            st_id_num, prev_quran, st_mobile, st_whatsapp, health_status,
            father_status, mother_status, kinship, p_id_num, whatsapp_num, wallet_num,
            bank_acc, bank_name, orig_addr, orig_type, orig_status,
            curr_addr, curr_type, notes
        ))
        
        student_id = cursor.lastrowid
        
        # Student Username = Student Identity Number directly without prefix
        clean_st_username = re.sub(r'\D', '', st_id_num) if st_id_num else str(student_id)
        
        cursor.execute("SELECT Id FROM Users WHERE Username = ?;", (clean_st_username,))
        if cursor.fetchone():
            clean_st_username = f"{clean_st_username}_{student_id}"
            
        cursor.execute("""
            INSERT INTO Users (Username, PasswordHash, PlainPassword, Role, FullName, IsActive, StudentId)
            VALUES (?, 'AQAAAAIAAYagAAAAEP4qOslLTFhJOkkBoFWNLPRj6QTbtdjPKCf+rxQ/YWVpdIgXBcSHPb6td4ysp1WmUA==', '123456', 3, ?, 1, ?);
        """, (clean_st_username, st_name, student_id))
        
        inserted_count += 1

conn.commit()
print(f"SUCCESS: Successfully imported {inserted_count} students with all 26 columns and {len(family_groups)} parent accounts into database!")

conn.close()
